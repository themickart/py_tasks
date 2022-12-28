from multiprocessing import Pool
from _datetime import datetime
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from jinja2 import Environment, FileSystemLoader
import pdfkit
from openpyxl.reader.excel import load_workbook
import pandas as pd
import os
import multiprocessing
from itertools import repeat


class DataSet:
    """
    Класс для хранения списка вакансий.

    Attributes:
        file_name (str): Название файла
        vacancies_objects (list): Список вакансий
    """

    def __init__(self, file_name):
        """
        Конструктор для инициализация объекта DataSet, который создает поле для хранения списка вакансий

        Args:
             file_name (str): Название файла
        """
        self.salary_by_year = dict()
        self.vacancies_count_by_year = dict()
        self.salary_by_profession_name = dict()
        self.vacancies_count_by_profession_name = dict()
        self.salary_by_city = dict()
        self.vacancy_rate_by_city = dict()
        self.dict_lict = list()


class InputConnect:
    """Класс для ввода данных и формирования отчетности о вакансиях

    Args:
        params (tuple): Кортеж с названием файла и профессии
    """

    def __init__(self):
        """Конструктор для инициализации объекта InputConnect"""
        self.file_name, self.path_name, self.profession_name = InputConnect.get_params()

    @staticmethod
    def get_params():
        """Статический метод для ввода данные о вакансии
        :return: Кортеж с названием файла и профессии
        """
        file_name = input("Введите название файла: ")
        path_name = "vacancies"
        profession_name = input("Введите название профессии: ")
        return file_name, path_name, profession_name

    @staticmethod
    def print_data_dict(self, data: DataSet):
        """Вычисляет и печатает в консоль словари со статистикой по городам
        :param self: Объект класса InputConnect
        :param data: Объект класса DataSet
        """
        df = pd.read_csv(self.file_name)
        count = len(df)
        df["salary"] = df[["salary_from", "salary_to"]].mean(axis=1)
        df["count"] = df.groupby("area_name")["area_name"].transform("count")
        df_norm = df[df["count"] > 0.01 * count]
        df_area = df_norm.groupby("area_name", as_index=False)["salary"].mean().sort_values(by="salary", ascending=False)
        df_area["salary"] = df_area["salary"].apply(lambda x: int(x))
        df_area10 = df_area.head(10)
        data.salary_by_city = dict(zip(df_area10["area_name"], df_area10["salary"]))

        data.vacancy_rate_by_city = {k: round(v / count, 4) for k, v in dict(df["area_name"].value_counts()).items()}

        print(f"Динамика уровня зарплат по годам: {data.salary_by_year}")
        print(f"Динамика количества вакансий по годам: {data.vacancies_count_by_year}")
        print(f"Динамика уровня зарплат по годам для выбранной профессии: {data.salary_by_profession_name}")
        print(
            f"Динамика количества вакансий по годам для выбранной профессии: {data.vacancies_count_by_profession_name}")
        print(f"Уровень зарплат по городам (в порядке убывания): {data.salary_by_city}")
        print(f"Доля вакансий по городам (в порядке убывания): {dict(list(data.vacancy_rate_by_city.items())[:10])}")

    def read_csv_by_path(self, path: str):
        """Метод для многопоточной обработки csv файлов при помощи модуля pandas
        :param path: Путь до csv файла с вакансиями
        :param data: Объект класса DataSet
        """
        df = pd.read_csv(path)
        df["salary"] = df[["salary_from", "salary_to"]].mean(axis=1)
        df["published_at"] = df["published_at"].apply(lambda d: datetime(int(d[:4]), int(d[5:7]), int(d[8:10])).year)
        year = df["published_at"][0]
        df_vacancy = df["name"].str.contains(self.profession_name)

        filter_by_year = df["published_at"] == year
        salary_by_year = (year, int(df[filter_by_year]["salary"].mean()))
        vacancies_count_by_year = (year, len(df[filter_by_year]))
        salary_by_profession_name = (year, int(df[df_vacancy & filter_by_year]["salary"].mean()))
        vacancies_count_by_profession_name = (year, len(df[df_vacancy & filter_by_year]))
        return salary_by_year, vacancies_count_by_year, salary_by_profession_name, vacancies_count_by_profession_name

    def processesed(self, data: DataSet):
        """
        Метод, для создания потоков при помощи модуля multiprocessing и вывода на консоль статистики по годам
        :param data: Объект класса DataSet
        :return: None
        """
        process_args = [f"{self.path_name}/{file}" for file in os.listdir(self.path_name)]

        pool = Pool(processes=multiprocessing.cpu_count())
        # pool = Pool()
        # stat_list = list(pool.starmap(InputConnect.read_csv_by_path, zip(repeat(self), process_args)))
        for item in pool.map(self.read_csv_by_path, process_args):
            data.salary_by_year[item[0][0]] = item[0][1]
            data.vacancies_count_by_year[item[1][0]] = item[1][1]
            data.salary_by_profession_name[item[2][0]] = item[2][1]
            data.vacancies_count_by_profession_name[item[3][0]] = item[3][1]
        pool.terminate()

        InputConnect.print_data_dict(self, data)

class Report:
    """Класс для формирования отчетности в виде pdf, excel или png файла"""

    @staticmethod
    def generate_excel(profession_name, data: DataSet):
        """Метод для генерации excel файла по названию профессии, после запуска данного метода
        файл с расширением xlsx появится в локальной директории проекта.

        :param profession_name: Название профессии
        :return: None
        """

        def as_text(value):
            """Функция, которая преобразует входное значение в тип str
            :param value: Any
            :return: str или "" Если value is None
            """
            if value is None:
                return ""
            return str(value)

        def set_max_length(worksheet):
            """Устанавливает максимальную длинну колонки в таблицу
            :param worksheet: Рабочая область таблицы
            """
            for column_cells in worksheet.columns:
                length = max(len(as_text(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

        def set_format_percent(worksheet):
            """Устанавливает в 5 колонке формат отображения данных в виде процентов
            :param worksheet: Рабочая область таблицы
            """
            for i, column_cells in enumerate(worksheet.columns):
                if i == 4:
                    for cell in column_cells:
                        cell.number_format = FORMAT_PERCENTAGE_00

        def set_border_style(worksheet):
            """Устанавливает стиль границам заполненных ячеек
            :param worksheet: Рабочая область таблицы
            """
            for column_cells in worksheet.columns:
                for cell in column_cells:
                    bd = Side(style="thin", color="000000")
                    cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)

        def set_headers(headers, head_range):
            """Устанавливает в первый ряд заголовки колонок
            :param headers: Список заголовок
            :param head_range: Диапазон значений для заголовок
            :return:
            """
            for i, cell in enumerate(head_range):
                cell.value = headers[i]
                cell.font = Font(size=11, b=True)

        wb = Workbook()
        sheet_1 = wb.worksheets[0]
        sheet_1.title = "Статистика по годам"
        sheet_2 = wb.create_sheet("Статистика по городам")
        headers = ["Год", "Средняя зарплата", f"Средняя зарплата - {profession_name}",
                   "Количество вакансий", f"Количество вакансий - {profession_name}"]
        set_headers(headers, sheet_1['A1':'E1'][0])

        for key in data.salary_by_year:
            sheet_1.append([key, data.salary_by_year[key], data.salary_by_profession_name[key],
                            data.vacancies_count_by_year[key], data.vacancies_count_by_profession_name[key]])
        set_border_style(sheet_1)
        set_max_length(sheet_1)

        set_headers(["Город", "Уровень зарплат"], sheet_2['A1':'B1'][0])
        set_headers(["Город", "Доля вакансий"], sheet_2['D1':'E1'][0])
        sheet_2.column_dimensions['C'].width = 2
        city_keys = list(data.vacancy_rate_by_city.keys())
        for i, key in enumerate(data.salary_by_city.keys()):
            sheet_2.append([key, data.salary_by_city[key], None, city_keys[i], data.vacancy_rate_by_city[city_keys[i]]])

        for i, column_cells in enumerate(sheet_2.columns):
            for cell in column_cells:
                if i != 2:
                    bd = Side(style="thin", color="000000")
                    cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        set_format_percent(sheet_2)
        set_max_length(sheet_2)
        wb.save("report.xlsx")
        return

    @staticmethod
    def generate_image(profession_name, data: DataSet):
        """Метод для генерирования картинки по названию профессии с графиками
        после запуска данного метода файл с расширением .png появится в локальной директории проекта.
        :param profession_name: Название професии
        """

        def myfunc(item):
            """Фукнция, которая устанавливает символ \n в строку, если в ней имеет символ ' ' или '-'
            :param item: Строка
            """
            if item.__contains__(' '):
                return item[:item.index(' ')] + '\n' + item[item.index(' ') + 1:]
            elif item.__contains__('-'):
                return item[:item.index('-')] + '-\n' + item[item.index('-') + 1:]
            return item

        width = 0.3
        nums = np.arange(len(data.salary_by_year.keys()))
        dx1 = nums - width / 2
        dx2 = nums + width / 2

        fig = plt.figure()
        ax = fig.add_subplot(221)
        ax.set_title("Уровень зарплат по годам")
        ax.bar(dx1, data.salary_by_year.values(), width, label="средняя з/п")
        ax.bar(dx2, data.salary_by_profession_name.values(), width, label=f"з/п {profession_name.lower()}")
        ax.set_xticks(nums, data.salary_by_year.keys(), rotation="vertical")
        ax.legend(fontsize=8)
        ax.tick_params(axis="both", labelsize=8)
        ax.grid(True, axis='y')

        ax = fig.add_subplot(222)
        ax.set_title("Количество вакансии по годам")
        ax.bar(dx1, data.vacancies_count_by_year.values(), width, label="Количество вакансии")
        ax.bar(dx2, data.vacancies_count_by_profession_name.values(), width,
               label=f"Количество вакансии\n{profession_name.lower()}")
        ax.set_xticks(nums, data.salary_by_year.keys(), rotation="vertical")
        ax.legend(fontsize=8)
        ax.tick_params(axis="both", labelsize=8)
        ax.grid(True, axis='y')

        ax = fig.add_subplot(223)
        ax.set_title("Уровень зарплат по городам")
        cities = list(map(myfunc, tuple(data.salary_by_city.keys())))
        y_pos = np.arange(len(cities))
        ax.barh(y_pos, list(data.salary_by_city.values()), align='center')
        ax.set_yticks(y_pos, labels=cities)
        ax.invert_yaxis()
        ax.grid(True, axis='x')

        ax = fig.add_subplot(224)
        ax.set_title("Доля вакансии по годам")
        labels = list(dict(list(data.vacancy_rate_by_city.items())[:10]).keys())
        labels.insert(0, "Другие")
        vals = list(dict(list(data.vacancy_rate_by_city.items())[:10]).values())
        vals.insert(0, 1 - sum(list(dict(list(data.vacancy_rate_by_city.items())[:10]).values())))
        ax.pie(vals, labels=labels, startangle=0, textprops={"fontsize": 6})
        plt.tight_layout()
        fig.set_size_inches(9.5, 7.5)
        plt.savefig("graph.png", dpi=120)
        return

    @staticmethod
    def generate_pdf(profession_name, data: DataSet):
        """Метода для генерации отчетности с графиком и таблицами.
        После запуска данного метода файл с расширением .pdf появится в локальной директории проекта.

        :param profession_name: Название профессии
        """
        Report.generate_excel(profession_name, data)
        Report.generate_image(profession_name, data)
        name = profession_name
        image_file = "graph.png"
        book = load_workbook("report.xlsx")
        sheet_1 = book.active
        sheet_2 = book['Статистика по городам']
        for row in range(2, sheet_2.max_row + 1):
            for col in range(4, 6):
                if type(sheet_2.cell(row, col).value).__name__ == "float":
                    sheet_2.cell(row, col).value = str(round(sheet_2.cell(row, col).value * 100, 2)) + '%'

        options = {'enable-local-file-access': None}
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        pdf_template = template.render({'name': name, 'image_file': image_file, 'sheet_1': sheet_1, 'sheet_2': sheet_2})
        config = pdfkit.configuration(wkhtmltopdf=r'D:\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options=options)


if __name__ == '__main__':
    inputparam = InputConnect()
    start_time = datetime.now()
    dataset = DataSet(inputparam.file_name)
    inputparam.processesed(dataset)
    Report.generate_pdf(inputparam.profession_name, dataset)
    print(f"Total time: {datetime.now() - start_time}")