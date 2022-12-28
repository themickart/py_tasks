import csv
import math
from _datetime import datetime
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from jinja2 import Environment, FileSystemLoader
import pdfkit
from openpyxl.reader.excel import load_workbook
import doctest
from time import strptime

"""Словарь для перевода з/п в рубли"""
currency_to_rub = {
    "AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76,
    "KZT": 0.13, "RUR": 1, "UAH": 1.64, "USD": 60.66, "UZS": 0.0055}


class DataSet:
    """
    Класс для хранения списка вакансий.

    Attributes:
        file_name (str): Название файла
        vacancies_objects (list): Список вакансий
    """
    def __init__(self, file_name):
        """
        Конструктор для инициализация объекта DataSet, который создает поле для хранения списка вакансий

        Args:
             file_name (str): Название файла
        """
        self.file_name = file_name
        self.vacancies_objects = list()

    def get_year_from_time(self, published_at: str):
        return strptime(published_at[:4], "%Y").tm_year

    def get_year_from_datetime(self, published_at: str):
        return datetime(int(published_at[:4]), int(published_at[5:7]), int(published_at[8:10])).year

    def get_year_from_datetime_strptime(self, published_at: str):
        return datetime.strptime(published_at, "%Y-%m-%dT%H:%M:%S%z").year

    def get_dataset(self):
        """ Считывает и фильтрует csv файл и формирует из строк объекты типа Vacancy для хранения в списке

        :return:
            DataSet: Объект класса DataSet

        >>> dataset = DataSet("vacancies.csv").get_dataset()
        >>> type(dataset.vacancies_objects[0]).__name__
        'Vacancy'

        >>> dataset = DataSet("vacancies.csv").get_dataset()
        >>> len(dataset.vacancies_objects)
        91

        >>> dataset = DataSet("vacancies.csv").get_dataset()
        >>> dataset.vacancies_objects[0].salary_from
        80000.0

        >>> dataset = DataSet("vacancies.csv").get_dataset()
        >>> dataset.vacancies_objects[0].salary_to
        100000.0

        >>> dataset = DataSet("vacancies.csv").get_dataset()
        >>> dataset.vacancies_objects[0].published_at
        2022

        """
        data = self.csv_reader()
        for item in data[1]:
            vacancy = Vacancy([item[0], item[1], item[2], item[3], item[4], item[5]])
            vacancy.published_at = self.get_year_from_datetime(vacancy.published_at)
            self.vacancies_objects.append(vacancy)
        return self

    def csv_reader(self):
        """ Считывает csv файл

        :returns:
            str: Строка с заголовками csv файла
            list: Список строк

        >>> items = DataSet("vacancies.csv").csv_reader()
        >>> len(items[0])
        12

        >>> items = DataSet("vacancies.csv").csv_reader()
        >>> len(items[1])
        91

        >>> items = DataSet("vacancies.csv").csv_reader()
        >>> len(items[1][0][2])
        171

        """
        with open(self.file_name, "r", encoding="utf-8-sig", newline="") as file:
            data = [x for x in csv.reader(file)]
        columns = data[0]
        rows = [x for x in data[1:] if len(x) == len(columns) and not x.__contains__("")]
        return columns, rows

class Vacancy:
    """
    Класс для предоставления вакансии.

    Attributes:
        name (str): Название вакансии
        salary_from (float): Нижняя граница вилки оклада
        salary_to (float): Верхняя граница вилки оклада
        salary_currency (str): Валюта оклада
        area_name (str): Название региона
        published_at (str): Дата публикации вакансии

    >>> type(Vacancy(args=["Senior Python Developer", "4500", "5500", "EUR", "Санкт-Петербург", "2022"])).__name__
    'Vacancy'

    >>> Vacancy(args=["Senior Python Developer", "4500", "5500", "EUR", "Санкт-Петербург", "2022"]).name
    'Senior Python Developer'

    >>> Vacancy(args=["Senior Python Developer", "4500", "5500", "EUR", "Санкт-Петербург", "2022"]).salary_from
    4500.0

    >>> Vacancy(args=["Senior Python Developer", "4500", "5500", "EUR", "Санкт-Петербург", "2022"]).salary_to
    5500.0

    >>> Vacancy(args=["Senior Python Developer", "4500", "5500", "EUR", "Санкт-Петербург", "2022"]).salary_currency
    'EUR'
    """

    def __init__(self, args):
        """Конструктов для инициализации вакансии

        :param args: Список строк с данными о вакансии
        """
        self.name = args[0]
        self.salary_from = float(args[1])
        self.salary_to = float(args[2])
        self.salary_currency = args[3]
        self.area_name = args[4]
        self.published_at = args[5]


class InputConnect:
    """Класс для ввода данных и формирования отчетности о вакансиях

    Args:
        params (tuple): Кортеж с названием файла и профессии
    """
    def __init__(self):
        """Конструктор для инициализации объекта InputConnect"""
        self.params = InputConnect.get_params()

    @staticmethod
    def get_params():
        """Статический метод для ввода данные о вакансии
        :return: Кортеж с названием файла и профессии
        """
        file_name = "vacancies_by_year.csv" #input("Введите название файла: ")
        profession_name = "Программист" #input("Введите название профессии: ")
        return file_name, profession_name

    @staticmethod
    def print_data_dict(self, data: DataSet):
        """Вычисляет и печатает в консоль словари со статистикой о вакансиях
        :param self: Объект класса InputConnect
        :param data: Объект класса DataSet
        """
        def get_correct_vacancy_rate(data: DataSet):
            """Функция для подсчета средней з/п
            :param data: Объект класса DataSet
            :return: Отсортированный словарь со средней з/п
            """
            data.vacancy_rate_by_city = {x: round(y / len(data.vacancies_objects), 4) for x, y in
                                         data.vacancy_rate_by_city.items()}
            return dict(sorted(data.vacancy_rate_by_city.items(), key=lambda item: item[1], reverse=True))
        data.vacancies_count_by_year = InputConnect.get_vacancies_count_by_name(data, "None")
        data.salary_by_year = InputConnect.get_salary_by_name(data, "None")
        data.vacancies_count_by_profession_name = InputConnect.get_vacancies_count_by_name(data, self.params[1])
        data.salary_by_profession_name = InputConnect.get_salary_by_name(data, self.params[1])
        data.vacancy_rate_by_city = InputConnect.get_vacancy_rate_by_city(data)
        data.salary_by_city = InputConnect.get_salary_by_city(data)
        data.vacancy_rate_by_city = get_correct_vacancy_rate(data)
        data.dict_lict = [data.salary_by_year, data.salary_by_profession_name, data.vacancies_count_by_year,
                          data.vacancies_count_by_profession_name, dict(list(data.salary_by_city.items())[:10]),
                          data.vacancy_rate_by_city]
        print(f"Динамика уровня зарплат по годам: {data.salary_by_year}")
        print(f"Динамика количества вакансий по годам: {data.vacancies_count_by_year}")
        print(f"Динамика уровня зарплат по годам для выбранной профессии: {data.salary_by_profession_name}")
        print(f"Динамика количества вакансий по годам для выбранной профессии: {data.vacancies_count_by_profession_name}")
        print(f"Уровень зарплат по городам (в порядке убывания): {dict(list(data.salary_by_city.items())[:10])}")
        print(f"Доля вакансий по городам (в порядке убывания): {dict(list(data.vacancy_rate_by_city.items())[:10])}")

    @staticmethod
    def get_vacancies_count_by_name(data: DataSet, name):
        """Статический метод для посчета количества вакансии в определенный по названию профессии
        :param data: Объект класса DataSet
        :param name: Название профессии
        :return: Словарь, где key - год, а value - количество профессии в определенный год

        >>> dataset = DataSet("vacancies.csv").get_dataset()
        >>> vacancies_count = InputConnect.get_vacancies_count_by_name(dataset, "Аналитик")
        >>> len(vacancies_count)
        1

        >>> dataset = DataSet("vacancies.csv").get_dataset()
        >>> vacancies_count = InputConnect.get_vacancies_count_by_name(dataset, "Аналитик")
        >>> vacancies_count[2022]
        2

        >>> dataset = DataSet("vacancies.csv").get_dataset()
        >>> vacancies_count = InputConnect.get_vacancies_count_by_name(dataset, "None")
        >>> len(vacancies_count)
        1

        >>> dataset = DataSet("vacancies.csv").get_dataset()
        >>> vacancies_count = InputConnect.get_vacancies_count_by_name(dataset, "None")
        >>> vacancies_count[2022]
        91

        """
        vacancies_count = dict()
        for vacancy in data.vacancies_objects:
            if vacancy.name.__contains__(name) or name == "None":
                InputConnect.set_value_by_name(vacancies_count, vacancy.published_at)
        if len(vacancies_count) == 0:
            return {2022: 0}
        return vacancies_count

    @staticmethod
    def get_salary_by_name(data: DataSet, name):
        """Статический метод для подсчета з/п по названию профессии
        :param data: Объект класса DataSet
        :param name: Название профессии
        :return: Словарь, где key - год, а value - средний оклад з/п в определенный год
        """
        salary_by_name = dict()
        for vacancy in data.vacancies_objects:
            if vacancy.name.__contains__(name) or name == "None":
                if not salary_by_name.__contains__(vacancy.published_at):
                    salary_by_name[vacancy.published_at] = InputConnect.get_currency_to_rub(vacancy)
                else:
                    salary_by_name[vacancy.published_at] += InputConnect.get_currency_to_rub(vacancy)
        if len(salary_by_name) == 0:
            return {2022: 0}
        for key in salary_by_name.keys():
            try:
                if name == "None":
                    salary_by_name[key] = math.floor(salary_by_name[key] / data.vacancies_count_by_year[key])
                else:
                    salary_by_name[key] = math.floor(salary_by_name[key] / data.vacancies_count_by_profession_name[key])
            except (ZeroDivisionError):
                raise ZeroDivisionError("Словарь 'vacancies_count_by_year' или 'vacancies_count_by_profession_name'"
                                        "содержит в качестве значение 0")
        return salary_by_name

    @staticmethod
    def get_vacancy_rate_by_city(data: DataSet):
        """Статический метод для подсчета количества общее вакансии в определенном регионе
        :param data: Объект класса DataSet
        :return: Словарь, где key - название региона, а value - общее количество вакансии в данном регионе

        >>> dataset = DataSet("vacancies.csv").get_dataset()
        >>> vacancy_rate = InputConnect.get_vacancy_rate_by_city(dataset)
        >>> len(vacancy_rate)
        36

        >>> dataset = DataSet("vacancies.csv").get_dataset()
        >>> vacancy_rate = InputConnect.get_vacancy_rate_by_city(dataset)
        >>> vacancy_rate["Москва"]
        24

        >>> dataset = DataSet("vacancies.csv").get_dataset()
        >>> vacancy_rate = InputConnect.get_vacancy_rate_by_city(dataset)
        >>> vacancy_rate["Санкт-Петербург"]
        12

        >>> dataset = DataSet("vacancies.csv").get_dataset()
        >>> vacancy_rate = InputConnect.get_vacancy_rate_by_city(dataset)
        >>> vacancy_rate["Екатеринбург"]
        3
        """
        vacancy_rate = dict()
        for vacancy in data.vacancies_objects:
            InputConnect.set_value_by_name(vacancy_rate, vacancy.area_name)
        return vacancy_rate

    @staticmethod
    def get_salary_by_city(data: DataSet):
        """Статический метод для подсчета средней з/п в определенном регионе при условии, что в данном регионе
        процент вакансии больше чем 1.
        :param data: Объект класса DataSet
        :return: Отсортированный словарь, где key - название региона, а value - средняя з/п в данном регионе

        >>> dataset = DataSet("vacancies.csv").get_dataset()
        >>> dataset.vacancy_rate_by_city = InputConnect.get_vacancy_rate_by_city(dataset)
        >>> salary_by_city = InputConnect.get_salary_by_city(dataset)
        >>> len(salary_by_city)
        36

        >>> dataset = DataSet("vacancies.csv").get_dataset()
        >>> dataset.vacancy_rate_by_city = InputConnect.get_vacancy_rate_by_city(dataset)
        >>> salary_by_city = InputConnect.get_salary_by_city(dataset)
        >>> salary_by_city["Москва"]
        157438

        >>> dataset = DataSet("vacancies.csv").get_dataset()
        >>> dataset.vacancy_rate_by_city = InputConnect.get_vacancy_rate_by_city(dataset)
        >>> salary_by_city = InputConnect.get_salary_by_city(dataset)
        >>> salary_by_city["Сочи"]
        26000

        >>> dataset = DataSet("vacancies.csv").get_dataset()
        >>> dataset.vacancy_rate_by_city = InputConnect.get_vacancy_rate_by_city(dataset)
        >>> salary_by_city = InputConnect.get_salary_by_city(dataset)
        >>> salary_by_city["Екатеринбург"]
        103333

        """
        salary_by_city = dict()
        for vacancy in data.vacancies_objects:
            if math.floor(data.vacancy_rate_by_city[vacancy.area_name] / len(data.vacancies_objects) * 100) >= 1:
                if not salary_by_city.__contains__(vacancy.area_name):
                    salary_by_city[vacancy.area_name] = InputConnect.get_currency_to_rub(vacancy)
                else:
                    salary_by_city[vacancy.area_name] += InputConnect.get_currency_to_rub(vacancy)
        for key in salary_by_city:
            salary_by_city[key] = math.floor(salary_by_city[key] / data.vacancy_rate_by_city[key])
        return dict(sorted(salary_by_city.items(), key=lambda item: item[1], reverse=True))

    @staticmethod
    def set_value_by_name(vacancy_dict: dict, name):
        """Инкрементирует значение в словаре, при условии существования ключа
        :param vacancy_dict: Словарь со значениями
        :param name: Название региона
        """
        if not vacancy_dict.__contains__(name):
            vacancy_dict[name] = 1
        else:
            vacancy_dict[name] += 1

    @staticmethod
    def get_currency_to_rub(vacancy):
        """Вычисляет среднюю з/п вилки и переподит в рубли, при помощи словаря - currency_to_rub
        :param vacancy: Объект класса Vacancy
        :return: Средняя з/п вилки
        """
        course = currency_to_rub[vacancy.salary_currency]
        return int((vacancy.salary_from * course + vacancy.salary_to * course) / 2)


class Report:
    """Класс для формирования отчетности в виде pdf, excel или png файла

    Args:
        data (list): Список словарей со статистикой о вакансиях
    """
    def __init__(self, dict_lict: list()):
        """Конструктор для инициализации объекта Report
        :param dict_lict: Список словарей со статистикой о вакансиях
        """
        self.data = dict_lict

    def generate_excel(self, profession_name):
        """Метод для генерации excel файла по названию профессии, после запуска данного метода
        файл с расширением xlsx появится в локальной директории проекта.

        :param profession_name: Название профессии
        :return: None
        """
        def as_text(value):
            """Функция, которая преобразует входное значение в тип str
            :param value: Any
            :return: str или "" Если value is None
            """
            if value is None:
                return ""
            return str(value)

        def set_max_length(worksheet):
            """Устанавливает максимальную длинну колонки в таблицу
            :param worksheet: Рабочая область таблицы
            """
            for column_cells in worksheet.columns:
                length = max(len(as_text(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

        def set_format_percent(worksheet):
            """Устанавливает в 5 колонке формат отображения данных в виде процентов
            :param worksheet: Рабочая область таблицы
            """
            for i, column_cells in enumerate(worksheet.columns):
                if i == 4:
                    for cell in column_cells:
                        cell.number_format = FORMAT_PERCENTAGE_00

        def set_border_style(worksheet):
            """Устанавливает стиль границам заполненных ячеек
            :param worksheet: Рабочая область таблицы
            """
            for column_cells in worksheet.columns:
                for cell in column_cells:
                    bd = Side(style="thin", color="000000")
                    cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)

        def set_headers(headers, head_range):
            """Устанавливает в первый ряд заголовки колонок
            :param headers: Список заголовок
            :param head_range: Диапазон значений для заголовок
            :return:
            """
            for i, cell in enumerate(head_range):
                cell.value = headers[i]
                cell.font = Font(size=11, b=True)

        wb = Workbook()
        sheet_1 = wb.worksheets[0]
        sheet_1.title = "Статистика по годам"
        sheet_2 = wb.create_sheet("Статистика по городам")
        headers = ["Год", "Средняя зарплата", f"Средняя зарплата - {profession_name}",
                   "Количество вакансий", f"Количество вакансий - {profession_name}"]
        set_headers(headers, sheet_1['A1':'E1'][0])

        for key in self.data[0].keys():
            sheet_1.append([key, self.data[0][key], self.data[1][key], self.data[2][key], self.data[3][key]])
        set_border_style(sheet_1)
        set_max_length(sheet_1)

        set_headers(["Город", "Уровень зарплат"], sheet_2['A1':'B1'][0])
        set_headers(["Город", "Доля вакансий"], sheet_2['D1':'E1'][0])
        sheet_2.column_dimensions['C'].width = 2
        city_keys = list(self.data[5].keys())
        for i, key in enumerate(self.data[4].keys()):
            sheet_2.append([key, self.data[4][key], None, city_keys[i], self.data[5][city_keys[i]]])

        for i, column_cells in enumerate(sheet_2.columns):
            for cell in column_cells:
                if i != 2:
                    bd = Side(style="thin", color="000000")
                    cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        set_format_percent(sheet_2)
        set_max_length(sheet_2)
        wb.save("report.xlsx")
        return

    def generate_image(self, profession_name):
        """Метод для генерирования картинки по названию профессии с графиками
        после запуска данного метода файл с расширением .png появится в локальной директории проекта.
        :param profession_name: Название професии
        """
        def myfunc(item):
            """Фукнция, которая устанавливает символ \n в строку, если в ней имеет символ ' ' или '-'
            :param item: Строка
            """
            if item.__contains__(' '):
                return item[:item.index(' ')] + '\n' + item[item.index(' ')+1:]
            elif item.__contains__('-'):
                return item[:item.index('-')] + '-\n' + item[item.index('-') + 1:]
            return item

        width = 0.3
        nums = np.arange(len(self.data[0].keys()))
        dx1 = nums - width / 2
        dx2 = nums + width / 2

        fig = plt.figure()
        ax = fig.add_subplot(221)
        ax.set_title("Уровень зарплат по годам")
        ax.bar(dx1, self.data[0].values(), width, label="средняя з/п")
        ax.bar(dx2, self.data[1].values(), width, label=f"з/п {profession_name.lower()}")
        ax.set_xticks(nums, self.data[0].keys(), rotation="vertical")
        ax.legend(fontsize=8)
        ax.tick_params(axis="both", labelsize=8)
        ax.grid(True, axis='y')

        ax = fig.add_subplot(222)
        ax.set_title("Количество вакансии по годам")
        ax.bar(dx1, self.data[2].values(), width, label="Количество вакансии")
        ax.bar(dx2, self.data[3].values(), width, label=f"Количество вакансии\n{profession_name.lower()}")
        ax.set_xticks(nums, self.data[0].keys(), rotation="vertical")
        ax.legend(fontsize=8)
        ax.tick_params(axis="both", labelsize=8)
        ax.grid(True, axis='y')

        ax = fig.add_subplot(223)
        ax.set_title("Уровень зарплат по городам")
        cities = list(map(myfunc, tuple(self.data[4].keys())))
        y_pos = np.arange(len(cities))
        ax.barh(y_pos, list(self.data[4].values()), align='center')
        ax.set_yticks(y_pos, labels=cities)
        ax.invert_yaxis()
        ax.grid(True, axis='x')

        ax = fig.add_subplot(224)
        ax.set_title("Доля вакансии по годам")
        labels = list(dict(list(self.data[5].items())[:10]).keys())
        labels.insert(0, "Другие")
        vals = list(dict(list(self.data[5].items())[:10]).values())
        vals.insert(0, 1 - sum(list(dict(list(self.data[5].items())[:10]).values())))
        ax.pie(vals, labels=labels, startangle=0, textprops={"fontsize": 6})
        plt.tight_layout()
        fig.set_size_inches(9.5, 7.5)
        plt.savefig("graph.png", dpi=120)
        return

    def generate_pdf(self, profession_name):
        """Метода для генерации отчетности с графиком и таблицами.
        После запуска данного метода файл с расширением .pdf появится в локальной директории проекта.

        :param profession_name: Название профессии
        """
        self.generate_excel(profession_name)
        self.generate_image(profession_name)
        name = profession_name
        image_file = "graph.png"
        book = load_workbook("report.xlsx")
        sheet_1 = book.active
        sheet_2 = book['Статистика по городам']
        for row in range(2, sheet_2.max_row + 1):
            for col in range(4, 6):
                if type(sheet_2.cell(row, col).value).__name__ == "float":
                    sheet_2.cell(row, col).value = str(round(sheet_2.cell(row, col).value * 100, 2)) + '%'

        options = {'enable-local-file-access': None}
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        pdf_template = template.render({'name': name, 'image_file': image_file, 'sheet_1': sheet_1, 'sheet_2': sheet_2})
        config = pdfkit.configuration(wkhtmltopdf=r'D:\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options=options)


def main_pdf():
    """ Функция для запуска формирования отчета
    :return: None
    """

    inputparam = InputConnect()
    start_time = datetime.now()
    dataset = DataSet(inputparam.params[0]).get_dataset()
    InputConnect.print_data_dict(inputparam, dataset)
    report = Report(dataset.dict_lict)
    report.generate_pdf(inputparam.params[1])
    print(f"Total time: {datetime.now() - start_time}")


if __name__ == '__main__':
    doctest.testmod()