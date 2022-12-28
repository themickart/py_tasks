import math
from _datetime import datetime
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
from jinja2 import Environment, FileSystemLoader
import pdfkit
from openpyxl.reader.excel import load_workbook
import pandas as pd


class DataSet:
    """
    Класс для хранения списка вакансий.

    Attributes:
        file_name (str): Название файла
        vacancies_objects (list): Список вакансий
    """

    def __init__(self, file_name):
        """
        Конструктор для инициализация объекта DataSet, который создает поле для хранения списка вакансий

        Args:
             file_name (str): Название файла
        """
        self.file_name = file_name
        self.salary_by_year = dict()
        self.vacancies_count_by_year = dict()
        self.salary_by_profession_name = dict()
        self.vacancies_count_by_profession_name = dict()
        self.dict_lict = list()


class InputConnect:
    """Класс для ввода данных и формирования отчетности о вакансиях

    Args:
        params (tuple): Кортеж с названием файла и профессии
    """

    def __init__(self):
        """Конструктор для инициализации объекта InputConnect"""
        self.file_name, self.profession_name = InputConnect.get_params()

    @staticmethod
    def get_params():
        """Статический метод для ввода данные о вакансии
        :return: Кортеж с названием файла и профессии
        """
        file_name = input("Введите название файла: ")
        profession_name = input("Введите название профессии: ")
        return file_name, profession_name

    @staticmethod
    def print_data_dict(self, data: DataSet):
        """Вычисляет и печатает в консоль словари со статистикой о вакансиях
        :param self: Объект класса InputConnect
        :param data: Объект класса DataSet
        """
        df = pd.read_csv(data.file_name)
        df['salary'] = df['salary'].fillna(0)
        df['salary'] = df['salary'].astype("int64")
        df["published_at"] = df["published_at"].apply(lambda d: int(d[:4]))
        years = df["published_at"].unique()
        df_vacancy = df["name"].str.contains(self.profession_name)

        for year in years:
            filter_by_year = df["published_at"] == year
            data.salary_by_year[year] = int(df[filter_by_year]["salary"].mean())
            data.vacancies_count_by_year[year] = len(df[filter_by_year])
            data.salary_by_profession_name[year] = int(df[df_vacancy & filter_by_year]["salary"].mean())
            data.vacancies_count_by_profession_name[year] = len(df[df_vacancy & filter_by_year])


class Report:
    """Класс для формирования отчетности в виде pdf, excel или png файла"""

    @staticmethod
    def generate_excel(profession_name, data: DataSet):
        """Метод для генерации excel файла по названию профессии, после запуска данного метода
        файл с расширением xlsx появится в локальной директории проекта.

        :param profession_name: Название профессии
        :return: None
        """

        def as_text(value):
            """Функция, которая преобразует входное значение в тип str
            :param value: Any
            :return: str или "" Если value is None
            """
            if value is None:
                return ""
            return str(value)

        def set_max_length(worksheet):
            """Устанавливает максимальную длинну колонки в таблицу
            :param worksheet: Рабочая область таблицы
            """
            for column_cells in worksheet.columns:
                length = max(len(as_text(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

        def set_border_style(worksheet):
            """Устанавливает стиль границам заполненных ячеек
            :param worksheet: Рабочая область таблицы
            """
            for column_cells in worksheet.columns:
                for cell in column_cells:
                    bd = Side(style="thin", color="000000")
                    cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)

        def set_headers(headers, head_range):
            """Устанавливает в первый ряд заголовки колонок
            :param headers: Список заголовок
            :param head_range: Диапазон значений для заголовок
            :return:
            """
            for i, cell in enumerate(head_range):
                cell.value = headers[i]
                cell.font = Font(size=11, b=True)

        wb = Workbook()
        sheet_1 = wb.worksheets[0]
        sheet_1.title = "Статистика по годам"
        headers = ["Год", "Средняя зарплата", f"Средняя зарплата - {profession_name}",
                   "Количество вакансий", f"Количество вакансий - {profession_name}"]
        set_headers(headers, sheet_1['A1':'E1'][0])

        for key in data.salary_by_year:
            sheet_1.append([key, data.salary_by_year[key], data.salary_by_profession_name[key],
                            data.vacancies_count_by_year[key], data.vacancies_count_by_profession_name[key]])
        set_border_style(sheet_1)
        set_max_length(sheet_1)
        wb.save("report.xlsx")
        return

    @staticmethod
    def generate_image(profession_name, data: DataSet):
        """Метод для генерирования картинки по названию профессии с графиками
        после запуска данного метода файл с расширением .png появится в локальной директории проекта.
        :param profession_name: Название професии
        """
        width = 0.3
        nums = np.arange(len(data.salary_by_year.keys()))
        dx1 = nums - width / 2
        dx2 = nums + width / 2

        fig = plt.figure()
        ax = fig.add_subplot(211)
        ax.set_title("Уровень зарплат по годам")
        ax.bar(dx1, data.salary_by_year.values(), width, label="средняя з/п")
        ax.bar(dx2, data.salary_by_profession_name.values(), width, label=f"з/п {profession_name.lower()}")
        ax.set_xticks(nums, data.salary_by_year.keys(), rotation="vertical")
        ax.legend(fontsize=8)
        ax.tick_params(axis="both", labelsize=8)
        ax.grid(True, axis='y')

        ax = fig.add_subplot(212)
        ax.set_title("Количество вакансии по годам")
        ax.bar(dx1, data.vacancies_count_by_year.values(), width, label="Количество вакансии")
        ax.bar(dx2, data.vacancies_count_by_profession_name.values(), width,
               label=f"Количество вакансии\n{profession_name.lower()}")
        ax.set_xticks(nums, data.salary_by_year.keys(), rotation="vertical")
        ax.legend(fontsize=8)
        ax.tick_params(axis="both", labelsize=8)
        ax.grid(True, axis='y')
        plt.tight_layout()
        fig.set_size_inches(9.5, 7.5)
        plt.savefig("graph.png", dpi=120)
        return

    @staticmethod
    def generate_pdf(profession_name, data: DataSet):
        """Метода для генерации отчетности с графиком и таблицами.
        После запуска данного метода файл с расширением .pdf появится в локальной директории проекта.

        :param data: Объект класса DataSet со всеми заполнеными полями
        :param profession_name: Название профессии
        """
        Report.generate_excel(profession_name, data)
        Report.generate_image(profession_name, data)
        name = profession_name
        image_file = "graph.png"
        book = load_workbook("report.xlsx")
        sheet_1 = book.active
        options = {'enable-local-file-access': None}
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template_cropped.html")
        pdf_template = template.render({'name': name, 'image_file': image_file, 'sheet_1': sheet_1})
        config = pdfkit.configuration(wkhtmltopdf=r'D:\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report3_4_2.pdf', configuration=config, options=options)


inputparam = InputConnect()
dataset = DataSet(inputparam.file_name)
InputConnect.print_data_dict(inputparam, dataset)
Report.generate_pdf(inputparam.profession_name, dataset)