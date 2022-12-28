from unittest import TestCase, main
from Task2_1_3 import DataSet, Report, InputConnect
from openpyxl.reader.excel import load_workbook


class ReportTest(TestCase):
    def test_sheetnames_generate_excel(self):
        book = load_workbook("report.xlsx")
        self.assertEqual(len(book.sheetnames), 2)
        self.assertEqual(book.sheetnames[0], "Статистика по годам")
        self.assertEqual(book.sheetnames[1], "Статистика по городам")

    def test_len_row_and_column_first_sheet(self):
        self.create_excel_file()
        book = load_workbook("report.xlsx")
        sheet_1 = book.active
        self.assertEqual(sheet_1.max_column, 5)
        self.assertEqual(sheet_1.max_row, 2)

    def test_len_row_and_column_second_sheet(self):
        self.create_excel_file()
        book = load_workbook("report.xlsx")
        sheet_2 = book['Статистика по городам']
        self.assertEqual(sheet_2.max_column, 5)
        self.assertEqual(sheet_2.max_row, 11)

    def test_check_cell_value_in_first_sheet(self):
        self.create_excel_file()
        book = load_workbook("report.xlsx")
        sheet_1 = book.active
        expected_values = [2022, 94892, 101666, 91, 6]
        actual = []
        for row in range(2, sheet_1.max_row + 1):
            for col in range(1, sheet_1.max_column + 1):
                actual.append(sheet_1.cell(row, col).value)
        self.assertEqual(expected_values, actual)

    def test_check_cell_value_in_second_sheet(self):
        self.create_excel_file()
        book = load_workbook("report.xlsx")
        sheet_2 = book['Статистика по городам']
        expected_values = ["Москва", 157438, None, "Москва", 0.2637]
        actual = []
        for row in range(2, 3):
            for col in range(1, sheet_2.max_column + 1):
                actual.append(sheet_2.cell(row, col).value)
        self.assertEqual(expected_values, actual)

    @staticmethod
    def create_excel_file():
        inputparam = InputConnect()
        dataset = DataSet(inputparam.params[0]).get_dataset()
        InputConnect.print_data_dict(inputparam, dataset)
        report = Report(dataset.dict_lict)
        report.generate_excel(inputparam.params[1])


if __name__ == '__main__':
    main()